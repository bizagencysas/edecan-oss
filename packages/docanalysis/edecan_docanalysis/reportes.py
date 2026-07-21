"""Exportación de reportes a XLSX (`exportar_analisis`, ROADMAP_V2.md §7.7, §5).

Arma un `openpyxl.Workbook` con una hoja **"Resumen"** (título + encabezado y
texto de cada sección, en orden) y **una hoja adicional por cada sección que
traiga `tabla`** (encabezados en negrita en la fila 1, luego una fila por
elemento de `tabla.filas`). La hoja "Resumen" deja una línea `Tabla → ver
hoja '<nombre>'` apuntando a cada hoja de detalle. Pensado para juntar en un
solo archivo descargable lo que ya calcularon `analizar_tabla`/
`extraer_tablas_pdf`/`analizar_imagen`, no para volver a analizar nada.

Nombres de hoja: Excel prohíbe `: \\ / ? * [ ]` y limita a 31 caracteres —
`_nombre_hoja` sanea y trunca, y desambigua sufijando `" (2)"`, `" (3)"`, ...
si dos secciones normalizan al mismo nombre.

Límites duros: `_MAX_SECCIONES` secciones por reporte, `_MAX_FILAS_TABLA` /
`_MAX_COLUMNAS_TABLA` por tabla de sección (recorta en silencio el exceso;
documentado en `docs/analista.md`).
"""

from __future__ import annotations

import io
import re
from typing import Any

from edecan_core import Tool, ToolContext, ToolResult

from . import _s3
from ._util import slugify

_MAX_SECCIONES = 100
_MAX_FILAS_TABLA = 20_000
_MAX_COLUMNAS_TABLA = 200
_MAX_NOMBRE_HOJA = 31

_INVALIDOS_HOJA_RE = re.compile(r"[:\\/?*\[\]]")

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExportarAnalisisTool(Tool):
    name = "exportar_analisis"
    description = (
        "Exporta un análisis a un reporte XLSX descargable: una hoja 'Resumen' con "
        "el título y el texto de cada sección, y una hoja adicional por cada sección "
        "que traiga una tabla ({etiquetas, filas})."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "titulo": {"type": "string", "description": "Título del reporte."},
            "secciones": {
                "type": "array",
                "description": "Secciones del reporte, en el orden en que deben aparecer.",
                "items": {
                    "type": "object",
                    "properties": {
                        "encabezado": {"type": "string", "description": "Título de la sección."},
                        "texto": {
                            "type": "string",
                            "description": "Texto libre de la sección (opcional).",
                        },
                        "tabla": {
                            "type": "object",
                            "description": "Tabla opcional de la sección (va en hoja aparte).",
                            "properties": {
                                "etiquetas": {
                                    "type": "array",
                                    "items": {"type": "string"},
                                    "description": "Encabezados de columna.",
                                },
                                "filas": {
                                    "type": "array",
                                    "items": {"type": "array"},
                                    "description": "Filas de datos (lista de valores cada una).",
                                },
                            },
                            "required": ["etiquetas", "filas"],
                        },
                    },
                    "required": ["encabezado"],
                },
            },
        },
        "required": ["titulo", "secciones"],
    }

    async def run(self, ctx: ToolContext, args: dict[str, Any]) -> ToolResult:
        titulo = str(args.get("titulo") or "").strip() or "Reporte"
        secciones_arg = args.get("secciones")
        if not isinstance(secciones_arg, list) or not secciones_arg:
            return ToolResult(content="Necesito al menos una sección en 'secciones'.")
        if len(secciones_arg) > _MAX_SECCIONES:
            return ToolResult(
                content=(
                    f"Demasiadas secciones ({len(secciones_arg)}); el máximo es {_MAX_SECCIONES}."
                )
            )

        secciones: list[dict[str, Any]] = []
        for i, seccion_arg in enumerate(secciones_arg):
            seccion, error = _validar_seccion(seccion_arg, i)
            if error:
                return ToolResult(content=error)
            secciones.append(seccion)

        contenido = _construir_xlsx(titulo, secciones)

        filename = f"{slugify(titulo, default='reporte')}.xlsx"
        file_id = await _s3.subir_resultado(
            ctx, filename=filename, mime=_XLSX_MIME, contenido=contenido
        )
        n_tablas = sum(1 for s in secciones if s["tabla"] is not None)
        return ToolResult(
            content=(
                f"Exporté «{titulo}» a «{filename}»: {len(secciones)} sección(es), "
                f"{n_tablas} con tabla en hoja aparte."
            ),
            data={"file_id": str(file_id), "filename": filename, "mime": _XLSX_MIME},
        )


def _validar_seccion(seccion: Any, indice: int) -> tuple[dict[str, Any], str | None]:
    if not isinstance(seccion, dict):
        return {}, f"La sección #{indice + 1} debe ser un objeto."

    encabezado = str(seccion.get("encabezado") or "").strip() or f"Sección {indice + 1}"
    texto_arg = seccion.get("texto")
    texto = str(texto_arg).strip() if texto_arg is not None else ""

    tabla_arg = seccion.get("tabla")
    if tabla_arg is None:
        return {"encabezado": encabezado, "texto": texto, "tabla": None}, None

    if not isinstance(tabla_arg, dict):
        return {}, f"La sección «{encabezado}»: 'tabla' debe ser un objeto {{etiquetas, filas}}."

    etiquetas_arg = tabla_arg.get("etiquetas")
    filas_arg = tabla_arg.get("filas")
    if not isinstance(etiquetas_arg, list) or not etiquetas_arg:
        return {}, f"La sección «{encabezado}»: 'tabla.etiquetas' debe ser una lista no vacía."
    if not isinstance(filas_arg, list):
        return {}, f"La sección «{encabezado}»: 'tabla.filas' debe ser una lista de filas."

    etiquetas = [str(e) for e in etiquetas_arg[:_MAX_COLUMNAS_TABLA]]
    filas = [
        list(fila)[: len(etiquetas)] if isinstance(fila, list | tuple) else [str(fila)]
        for fila in filas_arg[:_MAX_FILAS_TABLA]
    ]
    return {
        "encabezado": encabezado,
        "texto": texto,
        "tabla": {"etiquetas": etiquetas, "filas": filas},
    }, None


def _nombre_hoja(base: str, usados: set[str]) -> str:
    """Sanea `base` a un nombre de hoja Excel válido (sin `: \\ / ? * [ ]`,
    ≤31 caracteres) y lo desambigua sufijando `" (2)"`, `" (3)"`, ... contra
    `usados` (comparación case-insensitive, igual que hace Excel)."""
    limpio = _INVALIDOS_HOJA_RE.sub("-", base).strip() or "Tabla"
    limpio = limpio[:_MAX_NOMBRE_HOJA]

    candidato = limpio
    sufijo = 2
    while candidato.lower() in usados:
        marca = f" ({sufijo})"
        candidato = f"{limpio[: _MAX_NOMBRE_HOJA - len(marca)]}{marca}"
        sufijo += 1
    usados.add(candidato.lower())
    return candidato


def _construir_xlsx(titulo: str, secciones: list[dict[str, Any]]) -> bytes:
    import openpyxl  # import perezoso: solo lo necesita esta función
    from openpyxl.styles import Font

    libro = openpyxl.Workbook()
    resumen = libro.active
    resumen.title = "Resumen"

    resumen.append([titulo])
    resumen.cell(row=1, column=1).font = Font(bold=True, size=14)
    resumen.append([])

    nombres_usados: set[str] = set()
    fila_actual = 3
    for seccion in secciones:
        resumen.cell(row=fila_actual, column=1, value=seccion["encabezado"]).font = Font(bold=True)
        fila_actual += 1

        if seccion["texto"]:
            for linea in seccion["texto"].splitlines():
                resumen.cell(row=fila_actual, column=1, value=linea)
                fila_actual += 1

        tabla = seccion["tabla"]
        if tabla is not None:
            nombre_hoja = _nombre_hoja(seccion["encabezado"], nombres_usados)
            resumen.cell(row=fila_actual, column=1, value=f"Tabla → ver hoja '{nombre_hoja}'")
            fila_actual += 1

            hoja = libro.create_sheet(title=nombre_hoja)
            hoja.append(tabla["etiquetas"])
            for columna in range(1, len(tabla["etiquetas"]) + 1):
                hoja.cell(row=1, column=columna).font = Font(bold=True)
            for fila in tabla["filas"]:
                hoja.append(fila)

        fila_actual += 1  # línea en blanco entre secciones

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
