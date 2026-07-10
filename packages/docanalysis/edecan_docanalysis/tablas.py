"""Estadística descriptiva de tablas (`analizar_tabla`, ROADMAP_V2.md §7.7, §5).

Soporta **CSV** (delimitado por comas, UTF-8/UTF-8 con BOM) y **XLSX** (primera
hoja / hoja activa, vía `openpyxl` en modo `read_only=True` — no carga el libro
completo en memoria). Por columna infiere el tipo `numerica`/`texto`: es
`numerica` solo si TODOS los valores no vacíos de la muestra parsean como
número (`int`/`float` nativos en XLSX, o `float(str)` en CSV); si hay una sola
celda no numérica, la columna entera se trata como `texto`. Calcula:

- **Numérica**: `count`, `media`, `mediana`, `min`, `max`, `std` (desviación
  estándar MUESTRAL, `statistics.stdev` — `0.0` si hay menos de 2 valores) y
  `nulos`.
- **Texto**: `nulos` y las `_TOP_CATEGORIAS` categorías más frecuentes
  (`top`, lista de `{valor, conteo}` ordenada por conteo descendente y luego
  alfabéticamente para desempatar de forma determinista).

**Outliers**: rango intercuartílico (IQR) con cuartiles por el método de
Tukey (mediana de cada mitad de la lista ordenada, excluyendo la mediana
global cuando `n` es impar — el método más simple de enseñar/verificar a
mano). Un valor es atípico si cae fuera de `[Q1 - 1.5·IQR, Q3 + 1.5·IQR]`.
Con menos de `_MIN_VALORES_IQR` valores numéricos no se reportan outliers: el
rango intercuartílico no es significativo con tan poca muestra.

**Límites duros** (evitan que un archivo gigante tumbe el proceso o que la
respuesta al modelo sea inmanejable — ver `docs/analista.md`):
`_MAX_FILAS` filas de datos leídas por archivo, `_MAX_COLUMNAS` columnas
consideradas, `_TOP_CATEGORIAS` categorías por columna de texto,
`_MAX_OUTLIERS_POR_COLUMNA` valores atípicos listados por columna,
`_MAX_COLUMNAS_RESUMEN` columnas detalladas en el texto de respuesta (el
`data` estructurado sí trae todas las columnas hasta `_MAX_COLUMNAS`).

Si viene `pregunta`, se le pasan las estadísticas YA CALCULADAS (nunca las
filas crudas — no caben ni hace falta) a `ctx.llm.complete("principal", ...)`,
respetando el downgrade de modelo por plan vía `ctx.extras["flags"]` — mismo
patrón que `GenerarContenidoTool` en `edecan_toolkit.contenido`.
"""

from __future__ import annotations

import csv
import io
import statistics
from typing import Any

from edecan_core import Tool, ToolContext, ToolResult
from edecan_llm.base import ChatMessage, CompletionRequest

from . import _s3
from ._util import parse_uuid, tenant_flags

_MAX_FILAS = 50_000
_MAX_COLUMNAS = 200
_MAX_COLUMNAS_RESUMEN = 30
_TOP_CATEGORIAS = 5
_MAX_OUTLIERS_POR_COLUMNA = 20
_MIN_VALORES_IQR = 4

_CSV_MIMES = {"text/csv", "application/csv", "application/vnd.ms-excel"}
_XLSX_MIMES = {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

_SYSTEM_PROMPT = (
    "Eres un analista de datos. Se te dan estadísticas YA CALCULADAS de una "
    "tabla (no los datos crudos) y una pregunta del usuario sobre ellas. "
    "Responde en español, directo, basándote SOLO en las estadísticas dadas; "
    "si no alcanzan para responder con certeza, dilo explícitamente en vez de "
    "inventar un número."
)


class AnalizarTablaTool(Tool):
    name = "analizar_tabla"
    description = (
        "Analiza un archivo CSV o XLSX ya subido: infiere el tipo de cada columna, "
        "calcula estadística descriptiva (media, mediana, min, max, desviación "
        "estándar, nulos) para columnas numéricas, las categorías más frecuentes "
        "para columnas de texto, y detecta valores atípicos (outliers) por IQR. "
        "Opcionalmente responde una pregunta en lenguaje natural sobre esas "
        "estadísticas."
    )
    input_schema = {
        "type": "object",
        "properties": {
            "file_id": {
                "type": "string",
                "description": "id del archivo ya subido (CSV o XLSX) a analizar.",
            },
            "pregunta": {
                "type": "string",
                "description": "Pregunta opcional en lenguaje natural sobre la tabla.",
            },
        },
        "required": ["file_id"],
    }

    async def run(self, ctx: ToolContext, args: dict[str, Any]) -> ToolResult:
        file_id = parse_uuid(args.get("file_id"))
        if file_id is None:
            return ToolResult(content="'file_id' no es un identificador válido.")

        archivo = await _s3.descargar_archivo(ctx, file_id)
        if archivo is None:
            return ToolResult(content="No encontré ese archivo.")

        formato = _detectar_formato(archivo.mime, archivo.filename)
        if formato is None:
            return ToolResult(
                content=(
                    f"'{archivo.filename}' no es un CSV ni un XLSX — "
                    "analizar_tabla solo soporta esos dos formatos."
                )
            )

        try:
            encabezados, filas, total_filas_archivo = _parsear(archivo.contenido, formato)
        except Exception as exc:  # noqa: BLE001 - archivo corrupto = error de negocio, no de programación
            return ToolResult(content=f"No pude leer '{archivo.filename}': {exc}")

        if not encabezados:
            return ToolResult(content=f"'{archivo.filename}' está vacío o no tiene encabezados.")

        encabezados = encabezados[:_MAX_COLUMNAS]
        filas = [fila[: len(encabezados)] for fila in filas]

        columnas, stats, outliers = _analizar_columnas(encabezados, filas)
        resumen = _resumen_texto(
            archivo.filename, encabezados, len(filas), total_filas_archivo, columnas, stats
        )

        pregunta = str(args.get("pregunta") or "").strip()
        if pregunta:
            respuesta_llm = await _responder_pregunta(ctx, pregunta, columnas, stats, outliers)
            resumen = f"{resumen}\n\n{respuesta_llm}"

        return ToolResult(
            content=resumen,
            data={"columnas": columnas, "stats": stats, "outliers": outliers},
        )


def _detectar_formato(mime: str, filename: str) -> str | None:
    mime_normalizado = (mime or "").split(";")[0].strip().lower()
    nombre = (filename or "").lower()
    if mime_normalizado in _CSV_MIMES or nombre.endswith(".csv"):
        return "csv"
    if mime_normalizado in _XLSX_MIMES or nombre.endswith(".xlsx"):
        return "xlsx"
    return None


def _parsear(
    contenido: bytes, formato: str, hoja: str | None = None
) -> tuple[list[str], list[list[Any]], int]:
    """Devuelve `(encabezados, filas_de_datos_capadas_a_MAX_FILAS, total_filas_en_el_archivo)`.

    `hoja` (solo aplica a XLSX; CSV no tiene el concepto) selecciona una hoja por nombre en
    vez de la hoja activa/primera — parámetro ADITIVO agregado para `analizar_tabla_bytes`
    (WP-V6-06, `docs/analista.md`): con `hoja=None` (default) el comportamiento es IDÉNTICO al
    de antes, así que `AnalizarTablaTool.run()` (que nunca lo pasa) no cambia en nada.
    """
    if formato == "csv":
        return _parsear_csv(contenido)
    return _parsear_xlsx(contenido, hoja)


def _parsear_csv(contenido: bytes) -> tuple[list[str], list[list[Any]], int]:
    texto = contenido.decode("utf-8-sig", errors="replace")
    lector = csv.reader(io.StringIO(texto))
    filas_iter = iter(lector)
    try:
        encabezados = [c.strip() for c in next(filas_iter)]
    except StopIteration:
        return [], [], 0

    filas: list[list[Any]] = []
    total = 0
    for fila in filas_iter:
        if not fila or all(not str(c).strip() for c in fila):
            continue
        total += 1
        if len(filas) < _MAX_FILAS:
            filas.append(list(fila))
    return encabezados, filas, total


def _parsear_xlsx(
    contenido: bytes, hoja_nombre: str | None = None
) -> tuple[list[str], list[list[Any]], int]:
    import openpyxl  # import perezoso: solo lo necesita este formato

    libro = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    try:
        if hoja_nombre:
            if hoja_nombre not in libro.sheetnames:
                raise ValueError(
                    f"La hoja «{hoja_nombre}» no existe en este archivo. Hojas disponibles: "
                    f"{', '.join(libro.sheetnames)}."
                )
            hoja = libro[hoja_nombre]
        else:
            hoja = libro.active
        if hoja is None:
            return [], [], 0
        filas_iter = hoja.iter_rows(values_only=True)
        try:
            encabezados_raw = next(filas_iter)
        except StopIteration:
            return [], [], 0
        encabezados = [str(c).strip() if c is not None else "" for c in encabezados_raw]

        filas: list[list[Any]] = []
        total = 0
        for fila in filas_iter:
            if fila is None or all(c is None for c in fila):
                continue
            total += 1
            if len(filas) < _MAX_FILAS:
                filas.append(list(fila))
        return encabezados, filas, total
    finally:
        libro.close()


def _to_number(valor: Any) -> float | None:
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        s = valor.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _es_vacio(valor: Any) -> bool:
    return valor is None or (isinstance(valor, str) and not valor.strip())


def _analizar_columnas(
    encabezados: list[str], filas: list[list[Any]]
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, Any]]:
    columnas: list[dict[str, Any]] = []
    stats: dict[str, Any] = {}
    outliers: dict[str, Any] = {}
    total_filas = len(filas)

    for idx, nombre in enumerate(encabezados):
        clave = nombre or f"columna_{idx + 1}"
        valores_crudos = [fila[idx] if idx < len(fila) else None for fila in filas]
        no_vacios = [v for v in valores_crudos if not _es_vacio(v)]
        nulos = total_filas - len(no_vacios)

        numeros = [n for v in no_vacios if (n := _to_number(v)) is not None]
        es_numerica = bool(no_vacios) and len(numeros) == len(no_vacios)

        if es_numerica:
            columnas.append({"nombre": clave, "tipo": "numerica"})
            stats[clave] = _stats_numericas(numeros, nulos)
            outliers[clave] = _outliers_iqr(numeros)
        else:
            columnas.append({"nombre": clave, "tipo": "texto"})
            stats[clave] = _stats_texto(no_vacios, nulos)

    return columnas, stats, outliers


def _stats_numericas(numeros: list[float], nulos: int) -> dict[str, Any]:
    return {
        "tipo": "numerica",
        "count": len(numeros),
        "media": _round(statistics.mean(numeros)),
        "mediana": _round(statistics.median(numeros)),
        "min": _round(min(numeros)),
        "max": _round(max(numeros)),
        "std": _round(statistics.stdev(numeros)) if len(numeros) >= 2 else 0.0,
        "nulos": nulos,
    }


def _stats_texto(valores: list[Any], nulos: int) -> dict[str, Any]:
    contador: dict[str, int] = {}
    for v in valores:
        clave = str(v)
        contador[clave] = contador.get(clave, 0) + 1
    top = sorted(contador.items(), key=lambda par: (-par[1], par[0]))[:_TOP_CATEGORIAS]
    return {
        "tipo": "texto",
        "nulos": nulos,
        "top": [{"valor": valor, "conteo": conteo} for valor, conteo in top],
    }


def _outliers_iqr(numeros: list[float]) -> dict[str, Any]:
    if len(numeros) < _MIN_VALORES_IQR:
        return {"conteo": 0, "valores": []}
    ordenados = sorted(numeros)
    n = len(ordenados)
    mitad = n // 2
    mitad_inferior = ordenados[:mitad]
    mitad_superior = ordenados[mitad + 1 :] if n % 2 else ordenados[mitad:]
    q1 = statistics.median(mitad_inferior)
    q3 = statistics.median(mitad_superior)
    iqr = q3 - q1
    piso = q1 - 1.5 * iqr
    techo = q3 + 1.5 * iqr
    atipicos = [v for v in ordenados if v < piso or v > techo]
    return {
        "conteo": len(atipicos),
        "valores": [_round(v) for v in atipicos[:_MAX_OUTLIERS_POR_COLUMNA]],
    }


def _round(valor: float) -> float:
    return round(valor, 6)


def _resumen_texto(
    filename: str,
    encabezados: list[str],
    filas_leidas: int,
    total_filas_archivo: int,
    columnas: list[dict[str, Any]],
    stats: dict[str, Any],
) -> str:
    lineas = [f"Analicé «{filename}»: {len(encabezados)} columnas, {filas_leidas} filas de datos."]
    if total_filas_archivo > filas_leidas:
        lineas.append(
            f"(El archivo tiene {total_filas_archivo} filas de datos; por límite de "
            f"tamaño solo se analizaron las primeras {filas_leidas}.)"
        )
    for col in columnas[:_MAX_COLUMNAS_RESUMEN]:
        nombre = col["nombre"]
        s = stats[nombre]
        if col["tipo"] == "numerica":
            lineas.append(
                f"- {nombre} (numérica): media={s['media']}, mediana={s['mediana']}, "
                f"min={s['min']}, max={s['max']}, std={s['std']}, nulos={s['nulos']}"
            )
        else:
            top_txt = ", ".join(f"{t['valor']} ({t['conteo']})" for t in s["top"])
            lineas.append(f"- {nombre} (texto): nulos={s['nulos']}, top={top_txt or 'sin datos'}")
    if len(columnas) > _MAX_COLUMNAS_RESUMEN:
        lineas.append(
            f"… y {len(columnas) - _MAX_COLUMNAS_RESUMEN} columnas más "
            "(el detalle completo va en el resultado estructurado)."
        )
    return "\n".join(lineas)


async def _responder_pregunta(
    ctx: ToolContext,
    pregunta: str,
    columnas: list[dict[str, Any]],
    stats: dict[str, Any],
    outliers: dict[str, Any],
) -> str:
    contexto_lineas = ["Estadísticas por columna:"]
    for col in columnas:
        nombre = col["nombre"]
        contexto_lineas.append(f"- {nombre} ({col['tipo']}): {stats[nombre]}")
        if outliers.get(nombre, {}).get("conteo"):
            contexto_lineas.append(f"  outliers: {outliers[nombre]}")
    contexto = "\n".join(contexto_lineas)

    respuesta = await ctx.llm.complete(
        "principal",
        tenant_flags(ctx),
        CompletionRequest(
            model="principal",
            system=_SYSTEM_PROMPT,
            messages=[ChatMessage(role="user", content=f"{contexto}\n\nPregunta: {pregunta}")],
            max_tokens=1024,
        ),
    )
    return respuesta.text.strip() or "No pude generar una respuesta para esa pregunta."


# ---------------------------------------------------------------------------
# Superficie pública pura (sin S3 ni LLM) — WP-V6-06, `docs/analista.md`
# "Pantalla Analista". Reutilizan los mismos helpers privados de arriba
# (`_detectar_formato`/`_parsear`/`_analizar_columnas`) sin cambiar su
# lógica: son la MISMA estadística que ya calcula `AnalizarTablaTool.run()`,
# solo que reciben `bytes` ya descargados en vez de un `file_id` + `ToolContext`,
# y nunca llaman al LLM (la parte de `pregunta` se queda exclusiva del chat).
# Pensadas para consumidores fuera del loop de un agente — como el router
# HTTP de solo lectura `apps/api/edecan_api/routers/analista.py`.
# ---------------------------------------------------------------------------


def extraer_columnas_bytes(
    contenido: bytes, *, mime: str = "", filename: str = "", hoja: str | None = None
) -> dict[str, Any]:
    """Encabezados + filas CRUDAS (sin agregar a estadística) de un CSV/XLSX ya descargado.

    Misma detección de formato/límites que `AnalizarTablaTool.run()`
    (`_detectar_formato`, `_MAX_FILAS`, `_MAX_COLUMNAS`) — `ValueError` con el mismo mensaje en
    español si el formato no es soportado, el archivo no parsea, o queda sin encabezados.
    `hoja` (solo XLSX) selecciona una hoja por nombre; `None` usa la activa/primera.

    Devuelve `{"encabezados": list[str], "filas": list[list[Any]], "filas_leidas": int,
    "total_filas_archivo": int}` — `filas` ya recortadas a `len(encabezados)` columnas por
    fila. Pensada para consumidores que necesitan los VALORES crudos por columna (p. ej. para
    armar un gráfico o una serie de tiempo a partir de dos columnas elegidas por nombre) en vez
    de la estadística agregada que arma `analizar_tabla_bytes`/`AnalizarTablaTool`.
    """
    formato = _detectar_formato(mime, filename)
    if formato is None:
        raise ValueError(
            f"'{filename or 'archivo'}' no es un CSV ni un XLSX — solo se soportan esos dos "
            "formatos."
        )
    try:
        encabezados, filas, total_filas_archivo = _parsear(contenido, formato, hoja)
    except ValueError:
        raise
    except Exception as exc:  # noqa: BLE001 - archivo corrupto = error de negocio, no de programación
        raise ValueError(f"No pude leer '{filename or 'archivo'}': {exc}") from exc

    if not encabezados:
        raise ValueError(f"'{filename or 'archivo'}' está vacío o no tiene encabezados.")

    encabezados = encabezados[:_MAX_COLUMNAS]
    filas = [fila[: len(encabezados)] for fila in filas]

    return {
        "encabezados": encabezados,
        "filas": filas,
        "filas_leidas": len(filas),
        "total_filas_archivo": total_filas_archivo,
    }


def analizar_tabla_bytes(
    contenido: bytes, *, mime: str = "", filename: str = "", hoja: str | None = None
) -> dict[str, Any]:
    """Versión pura (sin S3 ni LLM) de la estadística que calcula `AnalizarTablaTool`: mismo
    tipo de columna, media/mediana/min/max/std/nulos, top de categorías y outliers IQR — la
    forma exacta que ya devuelve `ToolResult.data` (`columnas`/`stats`/`outliers`), sin
    inventar una forma nueva. NO incluye la parte de `pregunta` (eso exige `ctx.llm`, exclusivo
    del chat — ver el docstring del módulo); esta función nunca llama a ningún LLM.

    Construida sobre `extraer_columnas_bytes` (mismos `ValueError` si el formato no es
    soportado o el archivo no parsea) + `_analizar_columnas` (sin cambios).

    Devuelve `{"columnas": [...], "stats": {...}, "outliers": {...}, "filas_leidas": int,
    "total_filas_archivo": int}`.
    """
    extraido = extraer_columnas_bytes(contenido, mime=mime, filename=filename, hoja=hoja)
    columnas, stats, outliers = _analizar_columnas(extraido["encabezados"], extraido["filas"])
    return {
        "columnas": columnas,
        "stats": stats,
        "outliers": outliers,
        "filas_leidas": extraido["filas_leidas"],
        "total_filas_archivo": extraido["total_filas_archivo"],
    }
