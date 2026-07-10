"""Tests de `edecan_docanalysis.reportes` (`exportar_analisis`).

Verifica el XLSX subido leyéndolo de vuelta con `openpyxl` (round-trip real,
no solo inspección de los bytes crudos).
"""

from __future__ import annotations

import io

from edecan_docanalysis.reportes import ExportarAnalisisTool, _nombre_hoja


def _cargar_workbook(contenido: bytes):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(contenido))


async def test_titulo_y_secciones_requeridos(make_ctx, fake_s3):
    resultado = await ExportarAnalisisTool().run(make_ctx(), {"titulo": "X", "secciones": []})
    assert "al menos una sección" in resultado.content
    assert fake_s3.subidas == []


async def test_seccion_invalida(make_ctx, fake_s3):
    resultado = await ExportarAnalisisTool().run(
        make_ctx(), {"titulo": "X", "secciones": ["no-es-un-objeto"]}
    )
    assert "debe ser un objeto" in resultado.content


async def test_tabla_sin_etiquetas_es_invalida(make_ctx, fake_s3):
    resultado = await ExportarAnalisisTool().run(
        make_ctx(),
        {
            "titulo": "X",
            "secciones": [{"encabezado": "Datos", "tabla": {"etiquetas": [], "filas": []}}],
        },
    )
    assert "'tabla.etiquetas' debe ser una lista no vacía" in resultado.content


async def test_demasiadas_secciones(make_ctx, fake_s3):
    secciones = [{"encabezado": f"S{i}"} for i in range(101)]
    resultado = await ExportarAnalisisTool().run(
        make_ctx(), {"titulo": "X", "secciones": secciones}
    )
    assert "Demasiadas secciones" in resultado.content


async def test_reporte_completo_resumen_y_hoja_por_tabla(make_ctx, fake_s3):
    resultado = await ExportarAnalisisTool().run(
        make_ctx(),
        {
            "titulo": "Reporte de ventas",
            "secciones": [
                {"encabezado": "Introducción", "texto": "Resumen del trimestre.\nSegunda línea."},
                {
                    "encabezado": "Ventas por región",
                    "texto": "Ver detalle abajo.",
                    "tabla": {
                        "etiquetas": ["región", "total"],
                        "filas": [["Norte", 100], ["Sur", 200]],
                    },
                },
            ],
        },
    )

    assert "2 sección(es)" in resultado.content
    assert "1 con tabla" in resultado.content
    assert resultado.data["filename"] == "reporte-de-ventas.xlsx"

    subida = fake_s3.subidas[0]
    assert subida["mime"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert subida["filename"] == "reporte-de-ventas.xlsx"

    libro = _cargar_workbook(subida["contenido"])
    assert libro.sheetnames[0] == "Resumen"
    assert "Ventas por región" in libro.sheetnames

    resumen = libro["Resumen"]
    valores_resumen = [c.value for c in resumen["A"] if c.value is not None]
    assert valores_resumen[0] == "Reporte de ventas"
    assert "Introducción" in valores_resumen
    assert "Resumen del trimestre." in valores_resumen
    assert "Segunda línea." in valores_resumen
    assert "Ventas por región" in valores_resumen
    assert any("Tabla → ver hoja" in str(v) for v in valores_resumen)

    hoja_tabla = libro["Ventas por región"]
    filas = list(hoja_tabla.iter_rows(values_only=True))
    assert filas == [("región", "total"), ("Norte", 100), ("Sur", 200)]


async def test_seccion_sin_tabla_no_crea_hoja_extra(make_ctx, fake_s3):
    resultado = await ExportarAnalisisTool().run(
        make_ctx(), {"titulo": "X", "secciones": [{"encabezado": "Solo texto", "texto": "hola"}]}
    )
    subida = fake_s3.subidas[0]
    libro = _cargar_workbook(subida["contenido"])
    assert libro.sheetnames == ["Resumen"]
    assert "0 con tabla" in resultado.content


def test_nombre_hoja_sanea_caracteres_invalidos_y_trunca():
    usados: set[str] = set()
    nombre = _nombre_hoja("Ventas: Q1/Q2 [borrador]?*", usados)
    # cada carácter inválido (: / [ ] ? *) se reemplaza 1:1 por "-" (sin
    # colapsar corridas), ver `_INVALIDOS_HOJA_RE`.
    assert nombre == "Ventas- Q1-Q2 -borrador---"
    assert len(nombre) <= 31


def test_nombre_hoja_desambigua_duplicados_case_insensitive():
    usados: set[str] = set()
    n1 = _nombre_hoja("Datos", usados)
    n2 = _nombre_hoja("datos", usados)
    n3 = _nombre_hoja("Datos", usados)
    assert n1 == "Datos"
    assert n2 == "datos (2)"
    assert n3 == "Datos (3)"


async def test_dos_secciones_con_el_mismo_encabezado_producen_hojas_distintas(make_ctx, fake_s3):
    await ExportarAnalisisTool().run(
        make_ctx(),
        {
            "titulo": "X",
            "secciones": [
                {"encabezado": "Datos", "tabla": {"etiquetas": ["a"], "filas": [[1]]}},
                {"encabezado": "Datos", "tabla": {"etiquetas": ["b"], "filas": [[2]]}},
            ],
        },
    )
    subida = fake_s3.subidas[0]
    libro = _cargar_workbook(subida["contenido"])
    assert libro.sheetnames == ["Resumen", "Datos", "Datos (2)"]
    assert list(libro["Datos"].iter_rows(values_only=True)) == [("a",), (1,)]
    assert list(libro["Datos (2)"].iter_rows(values_only=True)) == [("b",), (2,)]
