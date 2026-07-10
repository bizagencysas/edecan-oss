"""Tests de `edecan_docanalysis.tablas` (`analizar_tabla`)."""

from __future__ import annotations

import io
from uuid import uuid4

import edecan_docanalysis.tablas as tablas_modulo
import pytest
from edecan_docanalysis.tablas import (
    AnalizarTablaTool,
    analizar_tabla_bytes,
    extraer_columnas_bytes,
)

_CSV = (
    b"nombre,edad,ciudad\n"
    b"Ana,10,Bogota\n"
    b"Luis,20,Medellin\n"
    b"Carlos,30,Bogota\n"
    b"Marta,40,Cali\n"
    b"Pedro,,Bogota\n"
)


async def test_file_id_invalido_no_toca_s3(make_ctx, fake_s3):
    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": "no-es-un-uuid"})
    assert "identificador válido" in resultado.content
    assert fake_s3.subidas == []


async def test_archivo_no_encontrado(make_ctx, fake_s3):
    fake_s3.archivo = None
    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})
    assert "No encontré ese archivo" in resultado.content


async def test_formato_no_soportado(make_ctx, fake_s3, make_archivo):
    fake_s3.archivo = make_archivo(
        contenido=b"%PDF-1.4", filename="datos.pdf", mime="application/pdf"
    )
    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})
    assert "no es un CSV ni un XLSX" in resultado.content


async def test_csv_vacio(make_ctx, fake_s3, make_archivo):
    fake_s3.archivo = make_archivo(contenido=b"", filename="vacio.csv", mime="text/csv")
    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})
    assert "vacío" in resultado.content


async def test_csv_infiere_tipos_y_calcula_estadistica_numerica_verificada_a_mano(
    make_ctx, fake_s3, make_archivo
):
    fake_s3.archivo = make_archivo(contenido=_CSV, filename="personas.csv", mime="text/csv")

    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})

    assert resultado.data["columnas"] == [
        {"nombre": "nombre", "tipo": "texto"},
        {"nombre": "edad", "tipo": "numerica"},
        {"nombre": "ciudad", "tipo": "texto"},
    ]

    # edad = [10, 20, 30, 40] (1 nulo, la fila de Pedro) — a mano:
    # media = (10+20+30+40)/4 = 25; mediana = (20+30)/2 = 25; min=10; max=40
    # varianza muestral = ((15)^2+(5)^2+(5)^2+(15)^2)/(4-1) = 500/3 = 166.6667
    # std = sqrt(166.6667) = 12.909944...
    edad = resultado.data["stats"]["edad"]
    assert edad == {
        "tipo": "numerica",
        "count": 4,
        "media": 25.0,
        "mediana": 25.0,
        "min": 10.0,
        "max": 40.0,
        "std": 12.909944,
        "nulos": 1,
    }

    # ciudad: Bogota aparece 3 veces, Cali y Medellin 1 vez cada una;
    # desempate alfabético entre Cali y Medellin.
    ciudad = resultado.data["stats"]["ciudad"]
    assert ciudad["nulos"] == 0
    assert ciudad["top"] == [
        {"valor": "Bogota", "conteo": 3},
        {"valor": "Cali", "conteo": 1},
        {"valor": "Medellin", "conteo": 1},
    ]

    # Sin outliers: con solo 4 valores el rango [Q1-1.5·IQR, Q3+1.5·IQR] es
    # [-15, 65], todos los valores de edad caen dentro.
    assert resultado.data["outliers"]["edad"] == {"conteo": 0, "valores": []}

    assert "Analicé «personas.csv»" in resultado.content
    assert "5 filas" in resultado.content


async def test_xlsx_infiere_tipos_y_calcula_estadistica(make_ctx, fake_s3, make_archivo):
    import openpyxl

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["producto", "precio", "categoria"])
    hoja.append(["Mesa", 100, "Muebles"])
    hoja.append(["Silla", 50, "Muebles"])
    hoja.append(["Lampara", 30, "Decoracion"])
    buffer = io.BytesIO()
    libro.save(buffer)

    fake_s3.archivo = make_archivo(
        contenido=buffer.getvalue(),
        filename="ventas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})

    assert resultado.data["columnas"] == [
        {"nombre": "producto", "tipo": "texto"},
        {"nombre": "precio", "tipo": "numerica"},
        {"nombre": "categoria", "tipo": "texto"},
    ]
    # precio = [100, 50, 30]: media=60, mediana=50, min=30, max=100
    # varianza muestral = ((40)^2+(-10)^2+(-30)^2)/(3-1) = 2600/2 = 1300
    # std = sqrt(1300) = 36.0555127...
    precio = resultado.data["stats"]["precio"]
    assert precio["count"] == 3
    assert precio["media"] == 60.0
    assert precio["mediana"] == 50.0
    assert precio["min"] == 30.0
    assert precio["max"] == 100.0
    assert precio["std"] == 36.055513
    assert precio["nulos"] == 0


async def test_outliers_iqr_detecta_el_unico_valor_atipico(make_ctx, fake_s3, make_archivo):
    csv = "x\n" + "\n".join(str(v) for v in [10, 11, 12, 13, 14, 15, 16, 90]) + "\n"
    fake_s3.archivo = make_archivo(contenido=csv.encode(), filename="x.csv", mime="text/csv")

    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})

    # Q1=11.5, Q3=15.5 (mediana de cada mitad, método de Tukey), IQR=4 ->
    # rango válido [5.5, 21.5]; 90 queda fuera, el resto no.
    assert resultado.data["outliers"]["x"] == {"conteo": 1, "valores": [90.0]}


async def test_pregunta_llama_al_llm_con_estadisticas_como_contexto(
    make_ctx, fake_s3, make_archivo, make_llm
):
    fake_s3.archivo = make_archivo(contenido=_CSV, filename="personas.csv", mime="text/csv")
    llm = make_llm(texto="La edad promedio es 25.")
    ctx = make_ctx(llm=llm, extras={"flags": {"models.premium": False}})

    resultado = await AnalizarTablaTool().run(
        ctx, {"file_id": str(uuid4()), "pregunta": "¿Cuál es la edad promedio?"}
    )

    assert "La edad promedio es 25." in resultado.content
    assert len(llm.llamadas) == 1
    alias, flags, req = llm.llamadas[0]
    assert alias == "principal"
    assert flags == {"models.premium": False}
    assert "edad" in req.messages[0].content
    assert "¿Cuál es la edad promedio?" in req.messages[0].content


async def test_limite_de_filas_se_respeta_y_se_avisa_en_el_resumen(
    make_ctx, fake_s3, make_archivo, monkeypatch
):
    monkeypatch.setattr(tablas_modulo, "_MAX_FILAS", 2)
    csv = "n\n" + "\n".join(str(i) for i in range(5)) + "\n"  # 5 filas de datos
    fake_s3.archivo = make_archivo(contenido=csv.encode(), filename="n.csv", mime="text/csv")

    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})

    assert resultado.data["stats"]["n"]["count"] == 2
    assert "primeras 2" in resultado.content


# ---------------------------------------------------------------------------
# `extraer_columnas_bytes` / `analizar_tabla_bytes` — superficie pública pura
# (WP-V6-06, sin S3 ni ToolContext ni LLM).
# ---------------------------------------------------------------------------


def test_extraer_columnas_bytes_devuelve_encabezados_y_filas_crudas():
    resultado = extraer_columnas_bytes(_CSV, mime="text/csv", filename="personas.csv")
    assert resultado["encabezados"] == ["nombre", "edad", "ciudad"]
    assert resultado["filas_leidas"] == 5
    assert resultado["total_filas_archivo"] == 5
    # Filas SIN agregar: el valor crudo de "edad" para Ana sigue siendo el string "10".
    assert resultado["filas"][0] == ["Ana", "10", "Bogota"]


def test_extraer_columnas_bytes_formato_no_soportado_lanza_valueerror():
    with pytest.raises(ValueError, match="no es un CSV ni un XLSX"):
        extraer_columnas_bytes(b"%PDF-1.4", mime="application/pdf", filename="x.pdf")


def test_extraer_columnas_bytes_vacio_lanza_valueerror():
    with pytest.raises(ValueError, match="vacío"):
        extraer_columnas_bytes(b"", mime="text/csv", filename="vacio.csv")


async def test_analizar_tabla_bytes_produce_la_misma_forma_que_la_tool(
    make_ctx, fake_s3, make_archivo
):
    """`analizar_tabla_bytes` sobre los mismos bytes que ya usa la tool debe dar EXACTAMENTE
    el mismo `columnas`/`stats`/`outliers` que `ToolResult.data` — es la misma estadística,
    solo sin pasar por S3/`ToolContext`."""
    resultado_puro = analizar_tabla_bytes(_CSV, mime="text/csv", filename="personas.csv")

    fake_s3.archivo = make_archivo(contenido=_CSV, filename="personas.csv", mime="text/csv")
    resultado_tool = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})

    assert resultado_puro["columnas"] == resultado_tool.data["columnas"]
    assert resultado_puro["stats"] == resultado_tool.data["stats"]
    assert resultado_puro["outliers"] == resultado_tool.data["outliers"]
    assert resultado_puro["filas_leidas"] == 5
    assert resultado_puro["total_filas_archivo"] == 5


def test_analizar_tabla_bytes_nunca_llama_al_llm():
    """No recibe `ctx` en absoluto — a diferencia de `AnalizarTablaTool.run()`, no hay `pregunta`
    ni forma de que esta función toque un LLM."""
    import inspect

    firma = inspect.signature(analizar_tabla_bytes)
    assert "ctx" not in firma.parameters
    assert "pregunta" not in firma.parameters


def test_analizar_tabla_bytes_xlsx_hoja_por_nombre_selecciona_esa_hoja():
    import openpyxl

    libro = openpyxl.Workbook()
    activa = libro.active
    activa.title = "resumen"
    activa.append(["nota"])
    activa.append(["esto no es lo que quiero analizar"])

    otra = libro.create_sheet("ventas")
    otra.append(["producto", "precio"])
    otra.append(["Mesa", 100])
    otra.append(["Silla", 50])

    buffer = io.BytesIO()
    libro.save(buffer)
    contenido = buffer.getvalue()
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Sin `hoja`: usa la activa ("resumen"), UNA sola columna de texto.
    sin_hoja = analizar_tabla_bytes(contenido, mime=mime, filename="libro.xlsx")
    assert sin_hoja["columnas"] == [{"nombre": "nota", "tipo": "texto"}]

    # Con `hoja="ventas"`: selecciona la otra hoja por nombre.
    con_hoja = analizar_tabla_bytes(contenido, mime=mime, filename="libro.xlsx", hoja="ventas")
    assert con_hoja["columnas"] == [
        {"nombre": "producto", "tipo": "texto"},
        {"nombre": "precio", "tipo": "numerica"},
    ]
    assert con_hoja["stats"]["precio"]["media"] == 75.0


def test_analizar_tabla_bytes_hoja_inexistente_lanza_valueerror_claro():
    import openpyxl

    libro = openpyxl.Workbook()
    libro.active.append(["x"])
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(ValueError, match="La hoja «no-existe» no existe"):
        analizar_tabla_bytes(
            buffer.getvalue(),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="libro.xlsx",
            hoja="no-existe",
        )


async def test_analizar_tabla_tool_sin_hoja_sigue_usando_la_activa_verificando_no_hay_regresion(
    make_ctx, fake_s3, make_archivo
):
    """`AnalizarTablaTool.run()` nunca pasa `hoja` — confirma que el parámetro nuevo (opcional,
    default `None`) no le cambia el comportamiento en nada."""
    import openpyxl

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["producto", "precio"])
    hoja.append(["Mesa", 100])
    buffer = io.BytesIO()
    libro.save(buffer)

    fake_s3.archivo = make_archivo(
        contenido=buffer.getvalue(),
        filename="x.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resultado = await AnalizarTablaTool().run(make_ctx(), {"file_id": str(uuid4())})
    assert resultado.data["columnas"] == [
        {"nombre": "producto", "tipo": "texto"},
        {"nombre": "precio", "tipo": "numerica"},
    ]
